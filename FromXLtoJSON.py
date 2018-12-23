from openpyxl import load_workbook
from elasticsearch import Elasticsearch
import json
# -------------------------------------------
class Forcast:
    def __init__(self):
        self.wb = load_workbook(filename="C:\\Users\\sleep\\Desktop\\forcast.xlsx")
        self.ws = self.wb.active
        self.INDX = {'colum':[1 ,2 ,3 ,4 ,5], "row":2}
        self.korPos = {}

        self.es = Elasticsearch()
    '''
    row = 1 
    col = 1 ('A')
    '''
    def ReadXl(self):
        while True:
            # ==============================================================================
            t_stp1 = self.ws.cell(row=self.INDX["row"], column=self.INDX['colum'][0]).value  # 서울특별시
            t_stp2 = self.ws.cell(row=self.INDX["row"], column=self.INDX['colum'][1]).value  # 종로구
            t_stp3 = self.ws.cell(row=self.INDX["row"], column=self.INDX['colum'][2]).value  # 사직동
            t_stp4 = self.ws.cell(row=self.INDX["row"], column=self.INDX['colum'][3]).value  # 격자 X
            t_stp5 = self.ws.cell(row=self.INDX["row"], column=self.INDX['colum'][4]).value  # 격자 Y
            # ==============================================================================
            # case _01_01 ( 작업할 데이터가 없는 경우 )
            if not(t_stp1): break
            # case _01_02 ( 작업할 데이터가 있는 경우 )
            else:
                # case _02_01 ( 2단계와 3단계의 데이터가 있을 경우 )
                # 두개 데이터가 모두 있는 경우
                if t_stp2 and t_stp3:
                    # case _02_01_01
                    t = {t_stp2:{t_stp3:{"x":t_stp4, "y":t_stp5}}}
                    try:
                        self.korPos[t_stp1]
                    except:
                        self.korPos[t_stp1] = t
                    else:
                        try:
                            self.korPos[t_stp1][t_stp2]
                        except:
                            self.korPos[t_stp1][t_stp2] = t[t_stp2]
                        else:
                            self.korPos[t_stp1][t_stp2].update(t[t_stp2])
                else: # 데이터가 없는 경우
                    # 두개 데이터가 모두 없는 경우
                    if not (t_stp2) and not (t_stp3):
                        t = {"x": t_stp4, "y": t_stp5}
                        try:
                            self.korPos[t_stp1]
                        except: # 해당하는 데이터가 비어있는 경우
                            self.korPos[t_stp1] = t
                        else: # 해당 데이터가 있는 경우
                            # 그러면 중복이 되는지 확인하여야 한다.
                            # case
                            if t not in self.korPos[t_stp1].values():
                                self.korPos[t_stp1].update(t)
                    elif t_stp2 and not (t_stp3):
                        t = {t_stp2:{"x": t_stp4, "y": t_stp5}}
                        try:
                            self.korPos[t_stp1]
                        except: # 해당하는 데이터가 비어있는 경우
                            self.korPos[t_stp1] = t
                        else: # 해당 데이터가 있는 경우
                            # 그러면 중복이 되는지 확인하여야 한다.
                            # case
                            if t not in self.korPos[t_stp1].values():
                                self.korPos[t_stp1].update(t)


            self.INDX["row"] += 1

    def jsonCreate(self):
        with open("forcastInfo.json", "w") as outfile:
            json.dump(self.korPos, outfile, ensure_ascii=False, indent='\t')

        outfile.close()

    # 소멸자
    def __del__(self):
        self.wb.close()

def main():
    fnode = Forcast() # 객체 생성
    fnode.ReadXl()
    fnode.jsonCreate()
if __name__ == "__main__":
    main()